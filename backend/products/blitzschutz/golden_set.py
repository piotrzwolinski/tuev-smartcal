"""Golden validation set: Blitzschutz_StV.xlsx — Stadtwerke Augsburg Ausschreibung.

Lokalizacja: ~/Desktop/TUEV/Anlagenliste-LV-Preisblätter_WP Blitzschutz_StV (1).xlsx
Sheet: 'Preisblatt 1-Anlagenliste-LV' (325 rows × 34 cols)

Struktura:
- 316 Anlagen Blitzschutz w Augsburg (Stadtwerke + extern)
- 4 cenniki (Einheitspreise pro Anlage):
    Col 31: Klasse I & II — Vollprüfung
    Col 32: Klasse I & II — Sichtprüfung
    Col 33: Klasse III & IV — Vollprüfung
    Col 34: Klasse III & IV — Sichtprüfung
- Per Anlage wypełniona tylko jedna z 4 cen (match Schutzklasse + Vollprüfung)

Merkmale ekstraktowalne:
- Standort-Strasse, PLZ, Ort
- Bezeichnung technische Anlage (filter "äußerer Blitzschutz")
- Objektgröße m² (BGF) — opcjonalne
- Anzahl der Trennstellen (= anzahl_ableitungen)
- Blitzschutzklasse (1/2/3/4 → Schutzklasse I/II/III/IV)
- (Ableitungen, Gebäudeumfang — w 82 obiektach Pos. 03.2)

Ground truth dla nas: Vollprüfung-cena per Anlage (col 31 dla I/II, col 33 dla III/IV).
"""

from pathlib import Path
from typing import List, Tuple, Optional

from products.blitzschutz.merkmale import (
    BlitzschutzMerkmale,
    Schutzklasse,
    GebaeudeNutzung,
)


BLITZSCHUTZ_STV_PATH = (
    Path.home() / "Desktop" / "TUEV"
    / "Anlagenliste-LV-Preisblätter_WP Blitzschutz_StV (1).xlsx"
)


# Heurystyka: rozpoznanie Gebäudenutzung z nazwy Standort
_NUTZUNG_KEYWORDS = [
    (["schule", "gymnasium", "berufschul", "volksschul"], GebaeudeNutzung.SCHULE),
    (["kindergarten", "kindertagesstätte", "kita"], GebaeudeNutzung.SCHULE),  # closest match
    (["jugendzentrum"], GebaeudeNutzung.SCHULE),
    (["büro", "verwaltung", "kreativ"], GebaeudeNutzung.BUERO),
    (["depot", "deponie", "betriebshof", "werk", "gaswerk"], GebaeudeNutzung.INDUSTRIE),
    (["lager", "halle", "depot"], GebaeudeNutzung.LAGER),
    (["wohn", "haus"], GebaeudeNutzung.WOHNUNG),
    (["hotel", "gaststätt"], GebaeudeNutzung.HOTEL),
    (["museum", "burg", "kirche"], GebaeudeNutzung.MUSEUM),
    (["krankenhaus", "klinik", "arzt"], GebaeudeNutzung.KRANKENHAUS),
    (["garage", "parkhaus"], GebaeudeNutzung.GARAGE),
]


def _classify_nutzung(name: Optional[str]) -> GebaeudeNutzung:
    if not name:
        return GebaeudeNutzung.SONSTIGE
    name_lower = str(name).lower()
    for keywords, nutzung in _NUTZUNG_KEYWORDS:
        if any(k in name_lower for k in keywords):
            return nutzung
    return GebaeudeNutzung.SONSTIGE


def _to_schutzklasse(value) -> Optional[Schutzklasse]:
    if value is None:
        return None
    s = str(value).strip()
    mapping = {"1": Schutzklasse.I, "2": Schutzklasse.II, "3": Schutzklasse.III, "4": Schutzklasse.IV}
    return mapping.get(s)


def _is_blitzschutz_aussen(bezeichnung: Optional[str]) -> bool:
    """5 wariantów nazwy w Excel-data — wszystkie äußerer Blitzschutz."""
    if not bezeichnung:
        return False
    s = str(bezeichnung).lower()
    return any(t in s for t in ["äußer", "äuß", "blitzschutz"])


def load_blitzschutz_golden_set() -> List[Tuple[BlitzschutzMerkmale, float, str]]:
    """Wczytuje 316 Anlagen + ich Vollprüfung-cena.

    Returns: list[(merkmale, expected_vollpruefung_price, reference)]
    """
    if not BLITZSCHUTZ_STV_PATH.exists():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(BLITZSCHUTZ_STV_PATH, data_only=True)
    ws = wb["Preisblatt 1-Anlagenliste-LV"]

    out = []
    # Headers in row 7, data from row 8. Read with iter_rows for reliability.
    rows = list(ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True))
    for row_idx, row in enumerate(rows, start=8):
        # Pad row to 34 columns
        cells = list(row) + [None] * (34 - len(row))
        bezeichnung = cells[12]  # Col 13 (0-indexed = 12)
        if not _is_blitzschutz_aussen(bezeichnung):
            continue

        nutzung_text = cells[5]   # Col 6
        strasse = cells[7]        # Col 8
        plz = cells[8]            # Col 9
        ort = cells[9]            # Col 10
        bgf = cells[21]           # Col 22

        # Merkmal-Wert pairs (cols 23-30 = idx 22-29)
        anzahl_trennstellen = None
        klasse = None
        gebaeudeumfang = None
        for mc, wc in [(22, 23), (24, 25), (26, 27), (28, 29)]:
            m_name = cells[mc]
            v = cells[wc]
            if not m_name:
                continue
            m_lower = str(m_name).lower()
            if "trennstell" in m_lower or "ableitung" in m_lower:
                try:
                    if anzahl_trennstellen is None:
                        anzahl_trennstellen = int(v)
                except (ValueError, TypeError):
                    pass
            elif "blitzschutzklasse" in m_lower:
                klasse = _to_schutzklasse(v)
            elif "gebäudeumfang" in m_lower or "umfang" in m_lower:
                try:
                    gebaeudeumfang = float(v)
                except (ValueError, TypeError):
                    pass

        if not anzahl_trennstellen or not klasse:
            continue

        # Ground-truth Vollprüfung
        if klasse in (Schutzklasse.I, Schutzklasse.II):
            expected = cells[30]  # Col 31 — Klasse I & II Vollprüfung
        else:
            expected = cells[32]  # Col 33 — Klasse III & IV Vollprüfung

        if not isinstance(expected, (int, float)) or expected <= 0:
            continue

        merkmale = BlitzschutzMerkmale(
            nutzung=_classify_nutzung(nutzung_text),
            schutzklasse=klasse,
            anzahl_ableitungen=anzahl_trennstellen,
            adresse_strasse=str(strasse) if strasse else None,
            adresse_plz=str(plz) if plz else None,
            adresse_ort=str(ort) if ort else None,
            adresse_lat=48.3668, adresse_lon=10.8865,  # Augsburg
            gebaeudeumfang_m=gebaeudeumfang,
            vereinsmitglied=True,
            erstpruefung=False,
            eilzuschlag=False,
        )

        ref = f"R{row_idx} · {str(nutzung_text)[:25] if nutzung_text else '?'} · {anzahl_trennstellen} TS · K{klasse.value}"
        out.append((merkmale, float(expected), ref))

    wb.close()
    return out

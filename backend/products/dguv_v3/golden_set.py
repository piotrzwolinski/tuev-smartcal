"""Golden validation set: DGUV V3 from Gersthofen Ausschreibung.

Source: Pruefung Elektrische Anlagen_Stadt Gersthofen_Preise_2025-04-14.xlsm
Sheet: GAEB_Konverter_LV (1.393 rows, Verteiler-Ebene)

Strategy: aggregate UV-level entries per Anlage (OZ group "01.", "02." etc.)
→ count UV/HV positions, sum GB (Gesamtbetrag) as ground truth price.
"""

from pathlib import Path
from typing import List, Tuple

from products.dguv_v3.merkmale import (
    DGUVMerkmale,
    GebaeudeNutzungDGUV,
    Installationskategorie,
)


GERSTHOFEN_PATH = (
    Path(__file__).resolve().parent.parent.parent.parent
    / "input-files"
    / "Pruefung Elektrische Anlagen_Stadt Gersthofen_Preise_2025-04-14.xlsm"
)


def load_dguv_golden_set() -> List[Tuple[DGUVMerkmale, float, str]]:
    if not GERSTHOFEN_PATH.exists():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(GERSTHOFEN_PATH, data_only=True, read_only=True)
    ws = wb["GAEB_Konverter_LV"]

    # Parse Anlage groups: rows with Art="NG" and OZ like "01.", "02." (no sub-position)
    anlagen = {}  # oz_prefix → {name, gb, uv_count}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None] * (12 - len(list(row)))
        oz = str(cells[1]).strip() if cells[1] else ""
        art = str(cells[2]).strip() if cells[2] else ""
        kurztext = str(cells[5]).strip() if cells[5] else ""
        gb = cells[9]

        # Anlage-level group: OZ like "01." or "02." (single level)
        if art == "NG - Normalgruppe" and oz.count(".") == 1 and not oz.endswith(".   "):
            prefix = oz.rstrip(".")
            if isinstance(gb, (int, float)) and gb > 0:
                anlagen[prefix] = {
                    "name": kurztext.split("\n")[0][:60],
                    "gb": float(gb),
                    "uv_count": 0,
                }

        # UV-level sub-group: OZ like "01.01." → count UVs per Anlage
        elif art == "NG - Normalgruppe" and oz.count(".") == 2:
            prefix = oz.split(".")[0]
            if prefix in anlagen:
                anlagen[prefix]["uv_count"] += 1

    out = []
    for prefix, data in sorted(anlagen.items()):
        uv_count = max(1, data["uv_count"])
        merkmale = DGUVMerkmale(
            nutzung=GebaeudeNutzungDGUV.SONSTIGE,
            gesamtflaeche_m2=uv_count * 200,  # heuristic: ~200 m² per UV
            primary_installationskategorie=Installationskategorie.KAT_1,
            anzahl_verteilungen_uv=uv_count,
            vereinsmitglied=True,
        )
        ref = f"Gersthofen OZ{prefix} · {data['name']} · {uv_count} UV · {data['gb']:.0f}€"
        out.append((merkmale, data["gb"], ref))

    wb.close()
    return out

"""Golden validation set: RLT from 441_419_Versand.xlsx (13.894 rows).

Source: Stefan Pausch Batch 2 (22.04.2026)
EQ Art → variant mapping: "Lüft.Garage"/"Garage" → GARAGE, rest → HYGIENE
Stefan-Filter: Faktura > 0 AND eigene Stunden > 0
"""

from pathlib import Path
from typing import List, Tuple

from products.rlt.merkmale import RLTMerkmale, RLTVariant


VERSAND_PATH = Path.home() / "Downloads" / "441_419_Versand.xlsx"

_GARAGE_KEYWORDS = ["garage", "tiefgarage", "parkhaus", "stellplatz"]


def _is_garage(eq_art: str) -> bool:
    if not eq_art:
        return False
    return any(k in eq_art.lower() for k in _GARAGE_KEYWORDS)


def load_rlt_golden_set() -> List[Tuple[RLTMerkmale, float, str]]:
    if not VERSAND_PATH.exists():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(VERSAND_PATH, data_only=True, read_only=True)
    ws = wb.active

    out = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row) + [None] * (24 - len(list(row)))

        eq_art = str(cells[3]) if cells[3] else ""
        faktura = cells[12]
        stunden = cells[17]

        if not isinstance(faktura, (int, float)) or faktura <= 0:
            continue
        if not isinstance(stunden, (int, float)) or stunden <= 0:
            continue

        is_garage = _is_garage(eq_art)
        variant = RLTVariant.GARAGE if is_garage else RLTVariant.HYGIENE

        if is_garage:
            merkmale = RLTMerkmale(
                variant=variant,
                baurechtlich=True,
                vereinsmitglied=True,
            )
        else:
            merkmale = RLTMerkmale(
                variant=variant,
                anzahl_pruefbereiche_hyg=1,
                vereinsmitglied=True,
            )

        ref = f"R{row_idx} · {eq_art[:30]} · {faktura:.0f}€ · {stunden:.1f}h"
        out.append((merkmale, float(faktura), ref))

    wb.close()
    return out
